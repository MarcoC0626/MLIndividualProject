from transformers import pipeline
from bs4 import BeautifulSoup
from transformers import AutoTokenizer, AutoModelForSequenceClassification
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.stem import PorterStemmer
from functools import reduce
from scipy.special import softmax
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from fake_useragent import UserAgent
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.styles import PatternFill
from progressbar import progressbar
import nltk
import string
import re
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl.drawing.image
import openpyxl

nltk.download('stopwords')
nltk.download('punkt')
nltk.download('averaged_perceptron_tagger')
ps = PorterStemmer()
stop_words = set(stopwords.words('english'))

#selenium related
options = Options()
options.add_argument('--headless')
options.add_argument("--disable-notifications")
#URL = 'https://www.amazon.com/VGCUB-Backpack-Approved-Waterproof-Business/dp/B0B6F558KR/ref=sr_1_5?crid=3AR4A95CR2KZG&keywords=backpack&qid=1697965747&sprefix=ba%2Caps%2C293&sr=8-5'
rev = []
pages = 0
preprocess = ""
#models related
sentiment_classifier = pipeline("sentiment-analysis") #default
text_classifier = pipeline("text-classification", model="LiYuan/amazon-review-sentiment-analysis") #amazon sentiment analysis
MODEL = f"cardiffnlp/twitter-roberta-base-sentiment"
pipe = pipeline("text-classification", model=MODEL) #roberta
tokenizer = AutoTokenizer.from_pretrained(MODEL)
model = AutoModelForSequenceClassification.from_pretrained(MODEL)

default_dict = []
amz_dict = []
roberta_dict = []

def get_url_pages():
    global pages, preprocess
    pagestop = 0
    url = input("Please Input a Amazon product URL: \n")
    while pagestop == 0:
        pages = int(input("How many pages of reviews you want?(Each page contains 10 reviews) \n"))
        if pages <= 0:
            print("Invalid input. Try again. \n")
        else:
            pagestop = 1
    preprocess = input("Do you need preprocessing? (y/n) \n")
    return url

def get_soup(url):
    ua = UserAgent()
    user_agent = ua.random
    options.add_argument(f'--user-agent={user_agent}')
    driver = webdriver.Chrome(options=options, executable_path=ChromeDriverManager().install())
    driver.get(url)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    detected = False
    elements = soup.select('p')
    for element in elements:
        if "Sorry, we just need to make sure you're not a robot." in element.text.strip() or "Sorry, your passkey isn't working." in element.text.strip():
            print("Detected by Amazon. Retrying.")
            detected = True
            break
    if detected:
        return get_soup(url)
    else:
        return soup

def Extractreviews(soup):
    for j in soup.findAll("span", {'data-hook': "review-body"}):
        rev.append(j.text)
    return rev

def Scraping():
    global review_url
    print("Start scraping, please wait")
    for i in range(pages):
        print(f"Extracting Page {i+1}")
        Extractreviews(get_soup(review_url))
        review_url = re.sub(f"pageNumber={i + 1}", f"pageNumber={i + 2}", review_url)
    print(f"Scraped {len(rev)} reviews.")

def RmPunc():
    for i in range(len(rev)):
        rev[i] = "".join([char for char in rev[i] if char not in string.punctuation])
def RmStopWords():
    for i in range(len(rev)):
        word_tokens = word_tokenize(rev[i])
        rev[i] = [w for w in word_tokens if not w.lower() in stop_words]
        rev[i] = (" ").join(rev[i])
def Stemming():
    for i in range(len(rev)):
        word_tokens = word_tokenize(rev[i])
        rev[i] = reduce(lambda x, y: x + " " + ps.stem(y), word_tokens, "")
def Cleaning():
    global rev
    rev = [i.strip() for i in rev]
    RmPunc()
    RmStopWords()
    Stemming()

def default_analysis(str, o_str):
    result = sentiment_classifier(str)[0]
    scores_dict = {
        'Analysis': result['label'],
        'Score': result['score'],
        'Review': o_str
    }
    default_dict.append(scores_dict)

def amz_analysis(str, o_str):
    result = text_classifier(str)[0]
    scores_dict = {
        'Analysis': result['label'],
        'Score': result['score'],
        'Review': o_str
    }
    amz_dict.append(scores_dict)

def roberta_analysis(str, o_str):
    encoded_text = tokenizer(str, return_tensors='pt')
    output = model(**encoded_text)
    scores = output[0][0].detach().numpy()
    scores = softmax(scores)    #rescales into 0-1 probabilities
    if max(scores) == scores[0]:
        scores_dict = {  # add labels and values
            'Analysis': "NEGATIVE",
            'Negative_Prob': scores[0],
            'Neutral_Prob': scores[1],
            'Positive_Prob': scores[2],
            'Review': o_str
        }
    elif max(scores) == scores[1]:
        scores_dict = {
            'Analysis': "NEUTRAL",
            'Negative_Prob': scores[0],
            'Neutral_Prob': scores[1],
            'Positive_Prob': scores[2],
            'Review': o_str
        }
    else:
        scores_dict = {
            'Analysis': "POSITIVE",
            'Negative_Prob': scores[0],
            'Neutral_Prob': scores[1],
            'Positive_Prob': scores[2],
            'Review': o_str
        }
    roberta_dict.append(scores_dict)

def analysis(list):
    print("Analysing. Please Wait.\n")
    for i in progressbar(range((len(rev)))):
        default_analysis(rev[i], list[i])
        amz_analysis(rev[i], list[i])
        roberta_analysis(rev[i], list[i])

def writetofile():
    d_default = pd.DataFrame(default_dict)
    d_amz = pd.DataFrame(amz_dict)
    d_roberta = pd.DataFrame(roberta_dict)
    d_empty = pd.DataFrame([])
    with (pd.ExcelWriter('Analysis.xlsx') as writer):
        d_default.to_excel(writer, sheet_name="Default Analysis", index=True, index_label="CustomerNo")
        sheet1 = writer.sheets["Default Analysis"]
        for cell, in sheet1[f'B2:B{len(d_default)+1}']:
            value = d_default["Analysis"].iloc[cell.row-2]
            cell.fill = PatternFill("solid", start_color=("5cb800" if value == "POSITIVE" else "ff2800"))
        d_empty.to_excel(writer, sheet_name="Default Analysis Plot")
        img = openpyxl.drawing.image.Image('default.png')
        worksheet = writer.sheets["Default Analysis Plot"]
        worksheet.add_image(img)

        d_amz.to_excel(writer, sheet_name="Amazon S.A. Model Analysis", index=True, index_label="CustomerNo")
        sheet2 = writer.sheets["Amazon S.A. Model Analysis"]
        for cell, in sheet2[f'B2:B{len(d_amz)+1}']:
            value = int(d_amz["Analysis"].iloc[cell.row-2].split(" ")[0])
            if value > 3:
                color = "5cb800"
            elif value < 3:
                color = "ff2800"
            else:
                color = "ffff00"
            cell.fill = PatternFill("solid", start_color=color)
        d_empty.to_excel(writer, sheet_name="Amazon model Plot")
        img = openpyxl.drawing.image.Image('amz.png')
        worksheet = writer.sheets["Amazon model Plot"]
        worksheet.add_image(img)

        d_roberta.to_excel(writer, sheet_name="Roberta Model Analysis", index=True, index_label="CustomerNo")
        sheet3 = writer.sheets["Roberta Model Analysis"]
        for cell, in sheet3[f'B2:B{len(d_roberta)+1}']:
            value = d_roberta["Analysis"].iloc[cell.row-2]
            if value == "POSITIVE":
                color = "5cb800"
            elif value == "NEGATIVE":
                color = "ff2800"
            else:
                color = "ffff00"
            cell.fill = PatternFill("solid", start_color=color)
        d_empty.to_excel(writer, sheet_name="Roberta model Plot")
        img = openpyxl.drawing.image.Image('roberta.png')
        worksheet = writer.sheets["Roberta model Plot"]
        worksheet.add_image(img)

def default_plot():
    pos = 0
    neg = 0
    for i in range(len(default_dict)):
        if default_dict[i]["Analysis"] == "POSITIVE":
            pos += 1
        else:
            neg += 1
    keys = ["Positive", "Negative"]
    values = [pos, neg]
    plt.bar(keys, values, width=0.5)
    plt.xlabel("Analysis")
    plt.ylabel("Number of reviews")
    plt.title("Default Analysis")
    plt.savefig("default.png")
    plt.close()

def amz_plot():
    stars = [0]*6
    for i in range(len(amz_dict)):
        stars[int(amz_dict[i]["Analysis"].split(" ")[0])] += 1
    keys = ["1 star", "2 stars", "3 stars", "4 stars", "5 stars"]
    values = [stars[1], stars[2], stars[3], stars[4], stars[5]]
    plt.bar(keys, values, width=0.5)
    plt.xlabel("Analysis")
    plt.ylabel("Number of reviews")
    plt.title("Amazon model Analysis")
    plt.savefig("amz.png")
    plt.close()

def roberta_plot():
    pos = 0
    neg = 0
    neu = 0
    for i in range(len(roberta_dict)):
        if roberta_dict[i]["Analysis"] == "POSITIVE":
            pos += 1
        elif roberta_dict[i]["Analysis"] == "NEGATIVE":
            neg += 1
        else:
            neu += 1
    keys = ["Positive", "Neutral", "Negative"]
    values = [pos, neu, neg]
    plt.bar(keys, values, width=0.5)
    plt.xlabel("Analysis")
    plt.ylabel("Number of reviews")
    plt.title("Roberta model Analysis")
    plt.savefig("roberta.png")
    plt.close()

#Main Program Begins
URL = get_url_pages()
review_url = URL.replace("dp", "product-reviews") + "&pageNumber=1&sortBy=recent"
Scraping()
rev_clone = rev.copy()
if preprocess.lower() == "y":
    Cleaning()

analysis(rev_clone)
default_plot()
amz_plot()
roberta_plot()
writetofile()
print("Finished. Results have been output to file.")
